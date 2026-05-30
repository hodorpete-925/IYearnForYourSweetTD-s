"""Quick: dump the first 5 rows of the Trades sheet so we can see how the
columns are actually laid out."""
import openpyxl
from pathlib import Path

EXCEL = Path("I Year For Your Sweet TD's Transaction + Draft Tracking.xlsx")
wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
ws = wb["Trades"]
print(f"Sheet 'Trades': {ws.max_row} rows x {ws.max_column} cols\n")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
    print(f"Row {i}: {list(row)}")
