"""reconcile_excel_vs_db.py - Compare Excel-tracked trades against DB.

Reads:
  - Excel 'Trades' sheet (Pete's manual log)
  - DB via the unioned all_transactions / all_transaction_players views
    (includes synthetic trades, filters vetoed)
  - transaction_overrides (to reclassify commish-pushed drop+add as trades)
  - transaction_picks (for the pick/FAAB diff)

Outputs (in place):
  - diff_excel_only.csv       — Excel rows with no DB match
  - diff_db_only.csv          — DB events with no Excel match
  - diff_picks_and_faab.csv   — Pick / FAAB trades flagged
  - db_only_review.txt        — narrative per DB-only event

Matching: ±2 days, normalized player names (suffixes, punctuation stripped).
Run:  python reconcile_excel_vs_db.py
"""

import csv
import difflib
import re
import sqlite3
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

DB_PATH = Path(__file__).parent / "fantasy.db"
EXCEL_PATH = Path(__file__).parent / "I Year For Your Sweet TD's Transaction + Draft Tracking.xlsx"
TRADES_SHEET = "Trades"

OUT_EXCEL_ONLY = Path(__file__).parent / "diff_excel_only.csv"
OUT_DB_ONLY = Path(__file__).parent / "diff_db_only.csv"
OUT_PICKS_FAAB = Path(__file__).parent / "diff_picks_and_faab.csv"
OUT_NARRATIVE = Path(__file__).parent / "db_only_review.txt"

DATE_TOLERANCE_DAYS = 2
OFF_SEASON_TOLERANCE_DAYS = 35  # Excel logs trade-agreement date, DB logs execution
NAME_FUZZ_THRESHOLD = 0.75       # difflib ratio cutoff for spelling variants


# ---------- helpers ----------
def normalize_name(name):
    """Lowercase, strip suffixes/punctuation, collapse whitespace."""
    if not name:
        return ""
    s = str(name).lower().strip()
    # Strip Sr/Jr/II/III/IV at end
    s = re.sub(r"\b(jr|sr|ii|iii|iv|v)\b\.?\s*$", "", s).strip()
    s = re.sub(r"[^\w\s]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def parse_excel_date(v):
    """Accept datetime or string, return date (no time)."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except Exception:
        try:
            return datetime.strptime(str(v), "%m/%d/%Y").date()
        except Exception:
            return None


def is_pick_row(player_str):
    """True if the 'player' value is actually a draft round / FAAB amount.
    Catches 'Round 4', '3rd Round (Malco)', '4th Round', etc."""
    if not player_str:
        return False
    s = str(player_str).strip().lower()
    if re.match(r"^round\s+\d+", s):
        return True
    if re.match(r"^\d+(st|nd|rd|th)?\s+round", s):
        return True
    if "faab" in s:
        return True
    return False


# ---------- Excel side ----------
def read_excel_trades():
    """Return list of dicts: {date, player, new_owner, trade_type, raw_row}."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    if TRADES_SHEET not in wb.sheetnames:
        print(f"ERROR: '{TRADES_SHEET}' sheet not found. Available: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[TRADES_SHEET]
    headers = None
    rows = []
    for i, raw in enumerate(ws.iter_rows(values_only=True), start=1):
        if headers is None:
            headers = [str(h).strip() if h else "" for h in raw]
            continue
        if not any(raw):
            continue
        row = {headers[j]: raw[j] for j in range(min(len(headers), len(raw)))}
        # Standardize to a small set of fields we care about. Excel column
        # names vary; try several common forms.
        date = (row.get("Date & Time") or row.get("trade_date") or
                row.get("Trade Date") or row.get("Date") or row.get("date"))
        player = (row.get("Player Name") or row.get("player") or
                  row.get("Player") or row.get("Players"))
        new_owner = (row.get("Manager Name (N)") or row.get("new_owner") or
                     row.get("New Owner") or row.get("To Manager") or
                     row.get("Receiver"))
        trade_type = (row.get("Trade Type") or row.get("trade_type") or
                      row.get("Type") or "")
        d = parse_excel_date(date)
        if d is None:
            continue
        rows.append({
            "date": d,
            "player": (str(player).strip() if player else ""),
            "new_owner": (str(new_owner).strip() if new_owner else ""),
            "trade_type": str(trade_type).strip(),
            "is_pick": is_pick_row(player),
            "matched": False,
        })
    print(f"Excel: read {len(rows)} trade rows from '{TRADES_SHEET}'")
    return rows


# ---------- DB side ----------
def read_db_trades(conn):
    """Return list of dicts of effective trade player-movements after applying
    the override layer."""
    # Pre-load override map: {transaction_id: source_team_season_id}
    overrides = {}
    for r in conn.execute(
        "SELECT transaction_id, override_type, source_team_season_id "
        "FROM transaction_overrides WHERE override_type = 'trade_from'"
    ):
        overrides[r[0]] = r[2]

    # Pull all_transactions + all_transaction_players for trades AND for
    # overridden drop+add events.
    rows = []
    cur = conn.execute("""
        SELECT t.transaction_id, t.timestamp, t.event_type, t.is_synthetic,
               tp.player_id, p.player_name,
               tp.direction, tp.team_season_id, tp.counterparty_team_season_id,
               tp.source_type
        FROM all_transactions t
        JOIN all_transaction_players tp ON tp.transaction_id = t.transaction_id
        JOIN players p ON p.player_id = tp.player_id
        WHERE tp.direction = 'incoming'
    """)
    for r in cur:
        (tx_id, ts, event_type, is_synth, pid, pname,
         direction, dest_team, src_team, src_type) = r
        is_trade = (event_type == "trade" or src_type == "team")
        if not is_trade and tx_id in overrides:
            is_trade = True
            src_team = overrides[tx_id]
            src_type = "team"
        if not is_trade:
            continue
        # Resolve manager names
        dest_mgr = _team_mgr(conn, dest_team)
        src_mgr = _team_mgr(conn, src_team) if src_team else "?"
        date = ts[:10]
        rows.append({
            "tx_id": tx_id,
            "date": datetime.strptime(date, "%Y-%m-%d").date(),
            "timestamp": str(ts),
            "player_id": pid,
            "player": pname,
            "new_owner": dest_mgr,
            "old_owner": src_mgr,
            "source_table": "synthetic" if is_synth else "real",
            "is_override": tx_id in overrides,
            "matched": False,
        })
    print(f"DB: assembled {len(rows)} effective trade movements "
          f"(after overrides + synthetic)")
    return rows


_team_cache = {}
def _team_mgr(conn, team_season_id):
    if team_season_id in _team_cache:
        return _team_cache[team_season_id]
    if team_season_id is None:
        return "?"
    row = conn.execute(
        "SELECT m.full_name FROM teams t JOIN managers m ON m.manager_id = t.manager_id "
        "WHERE t.team_season_id = ?", (team_season_id,)
    ).fetchone()
    name = row[0] if row else f"team#{team_season_id}"
    _team_cache[team_season_id] = name
    return name


def read_db_picks(conn):
    """Pick trades from transaction_picks. Returns list of dicts."""
    rows = []
    for r in conn.execute("""
        SELECT t.transaction_id, t.timestamp, tp.draft_round,
               src.team_season_id AS src_team, dst.team_season_id AS dst_team,
               ms.full_name, md.full_name
        FROM transaction_picks tp
        JOIN transactions t ON t.transaction_id = tp.transaction_id
        JOIN teams src ON src.team_season_id = tp.source_team_season_id
        JOIN teams dst ON dst.team_season_id = tp.destination_team_season_id
        JOIN managers ms ON ms.manager_id = src.manager_id
        JOIN managers md ON md.manager_id = dst.manager_id
        WHERE t.status = 'successful'
    """):
        tx_id, ts, rnd, src_team, dst_team, src_mgr, dst_mgr = r
        rows.append({
            "tx_id": tx_id,
            "date": datetime.strptime(ts[:10], "%Y-%m-%d").date(),
            "round": rnd,
            "new_owner": dst_mgr,
            "old_owner": src_mgr,
            "matched": False,
        })
    print(f"DB: {len(rows)} pick trades in transaction_picks")
    return rows


# ---------- matching ----------
def _tolerance_for(er):
    """Off-season trades get a much wider date window because the Excel
    logs the agreement date while the DB logs the execution date."""
    if "off-season" in (er.get("trade_type") or "").lower():
        return OFF_SEASON_TOLERANCE_DAYS
    return DATE_TOLERANCE_DAYS


def _fuzzy_name_match(excel_name, db_name):
    """Normalized exact match OR fuzzy ratio above threshold."""
    a = normalize_name(excel_name)
    b = normalize_name(db_name)
    if a == b:
        return True
    if not a or not b:
        return False
    ratio = difflib.SequenceMatcher(None, a, b).ratio()
    return ratio >= NAME_FUZZ_THRESHOLD


def match_excel_to_db(excel_rows, db_rows, db_picks):
    """Tolerant match: per-row date tolerance, fuzzy name, same new_owner."""
    # Index DB picks by round + new_owner (exact match on these)
    pick_index = defaultdict(list)
    for r in db_picks:
        pick_index[(r["round"], r["new_owner"])].append(r)
    # Group DB player rows by new_owner only — we'll do fuzzy name within group
    db_by_owner = defaultdict(list)
    for r in db_rows:
        db_by_owner[r["new_owner"]].append(r)

    for er in excel_rows:
        tol = _tolerance_for(er)
        if er["is_pick"]:
            pl = er["player"].lower()
            m = re.match(r"round\s+(\d+)", pl) or re.match(r"^(\d+)(st|nd|rd|th)?\s+round", pl)
            if not m:
                continue
            rnd = int(m.group(1))
            for cand in pick_index[(rnd, er["new_owner"])]:
                if cand["matched"]:
                    continue
                if abs((cand["date"] - er["date"]).days) <= tol:
                    er["matched"] = True
                    cand["matched"] = True
                    break
        else:
            for cand in db_by_owner.get(er["new_owner"], []):
                if cand["matched"]:
                    continue
                if abs((cand["date"] - er["date"]).days) > tol:
                    continue
                if _fuzzy_name_match(er["player"], cand["player"]):
                    er["matched"] = True
                    cand["matched"] = True
                    break


# ---------- output ----------
def write_diffs(excel_rows, db_rows, db_picks, excel_max_date):
    excel_only = [r for r in excel_rows if not r["matched"] and not r["is_pick"]]
    db_only_all = [r for r in db_rows if not r["matched"]]
    db_only = [r for r in db_only_all if r["date"] <= excel_max_date]
    db_out_of_scope = [r for r in db_only_all if r["date"] > excel_max_date]
    pick_diffs = [r for r in excel_rows if not r["matched"] and r["is_pick"]]
    unmatched_db_picks_all = [r for r in db_picks if not r["matched"]]
    unmatched_db_picks = [r for r in unmatched_db_picks_all if r["date"] <= excel_max_date]
    db_pick_out_of_scope = [r for r in unmatched_db_picks_all if r["date"] > excel_max_date]

    with open(OUT_EXCEL_ONLY, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["trade_date", "player", "new_owner", "trade_type"])
        for r in sorted(excel_only, key=lambda x: (x["date"], x["player"])):
            w.writerow([r["date"].isoformat(), r["player"], r["new_owner"], r["trade_type"]])

    with open(OUT_DB_ONLY, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["trade_date", "trade_ts", "player", "new_owner", "source_table",
                    "txn_id", "is_override"])
        for r in sorted(db_only, key=lambda x: (x["date"], x["player"])):
            w.writerow([r["date"].isoformat(), r["timestamp"], r["player"],
                        r["new_owner"], r["source_table"], r["tx_id"], r["is_override"]])

    with open(OUT_PICKS_FAAB, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["trade_date", "player", "new_owner", "trade_type", "side"])
        for r in sorted(pick_diffs, key=lambda x: (x["date"], x["player"])):
            w.writerow([r["date"].isoformat(), r["player"], r["new_owner"],
                        r["trade_type"], "excel_only"])
        for r in sorted(unmatched_db_picks, key=lambda x: (x["date"], x["round"])):
            w.writerow([r["date"].isoformat(), f"Round {r['round']}", r["new_owner"],
                        "", "db_only"])

    # Narrative summary per DB-only event
    by_tx = defaultdict(list)
    for r in db_only:
        by_tx[r["tx_id"]].append(r)
    with open(OUT_NARRATIVE, "w", encoding="utf-8") as f:
        f.write(f"{len(db_only)} unmatched DB player-movements across "
                f"{len(by_tx)} trade events\n\n")
        for tx_id, mvs in sorted(by_tx.items(), key=lambda kv: kv[1][0]["date"]):
            first = mvs[0]
            f.write("=" * 80 + "\n")
            f.write(f"txn {tx_id}  |  {first['timestamp']}  |  {first['source_table']}"
                    + ("  [override]" if first['is_override'] else "") + "\n")
            for r in mvs:
                f.write(f"  {r['player']:<30} {r['old_owner']:<20} -> {r['new_owner']}\n")
            f.write("\n")

    return excel_only, db_only, pick_diffs, unmatched_db_picks, db_out_of_scope, db_pick_out_of_scope


# ---------- main ----------
def main():
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        sys.exit(1)
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON;")

    excel_rows = read_excel_trades()
    db_rows = read_db_trades(conn)
    db_picks = read_db_picks(conn)
    match_excel_to_db(excel_rows, db_rows, db_picks)
    excel_max_date = max((r["date"] for r in excel_rows), default=None)
    if excel_max_date is None:
        print("ERROR: no Excel rows; nothing to compare against.")
        sys.exit(1)
    print(f"Excel max date: {excel_max_date} (DB events after this counted as out-of-scope)")
    excel_only, db_only, pick_diffs, db_pick_only, db_oos, db_pick_oos = write_diffs(
        excel_rows, db_rows, db_picks, excel_max_date)

    print()
    print("=" * 60)
    print("RECONCILIATION SUMMARY")
    print("=" * 60)
    print(f"Excel rows:                       {len(excel_rows):>5}")
    print(f"  matched to DB:                  {sum(1 for r in excel_rows if r['matched']):>5}")
    print(f"  unmatched (excel_only):         {len(excel_only):>5}")
    print(f"  unmatched pick/FAAB:            {len(pick_diffs):>5}")
    print(f"DB trade movements:               {len(db_rows):>5}")
    print(f"  matched to Excel:               {sum(1 for r in db_rows if r['matched']):>5}")
    print(f"  unmatched (db_only):            {len(db_only):>5}")
    print(f"DB pick trades:                   {len(db_picks):>5}")
    print(f"  matched to Excel:               {sum(1 for r in db_picks if r['matched']):>5}")
    print(f"  unmatched in scope:             {len(db_pick_only):>5}")
    print(f"  out of scope (post-{excel_max_date}): {len(db_pick_oos):>5}")
    print()
    print(f"DB-only player movements:         {len(db_only) + len(db_oos):>5}")
    print(f"  in scope (need reconciliation): {len(db_only):>5}")
    print(f"  out of scope (post-{excel_max_date}): {len(db_oos):>5}")
    print()
    print("Files written:")
    print(f"  {OUT_EXCEL_ONLY.name}")
    print(f"  {OUT_DB_ONLY.name}")
    print(f"  {OUT_PICKS_FAAB.name}")
    print(f"  {OUT_NARRATIVE.name}")


if __name__ == "__main__":
    main()
